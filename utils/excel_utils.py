from datetime import datetime
from openpyxl import load_workbook


class ExcelUtils:
    """Utility class for Excel file operations."""

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active

    def get_test_data(self):
        """Read all test data from Excel file."""
        test_data = []
        headers = [cell.value for cell in self.ws[1]]

        for row in self.ws.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for idx, value in enumerate(row):
                row_data[headers[idx]] = value
            test_data.append(row_data)

        return test_data

    def write_result(self, row_num, test_result):
        """Write test result to the specified row."""
        date_col = self.find_column_index("Date")
        time_col = self.find_column_index("Time of Test")
        result_col = self.find_column_index("Test Result")

        current_datetime = datetime.now()
        self.ws.cell(row=row_num, column=date_col, value=current_datetime.strftime("%Y-%m-%d"))
        self.ws.cell(row=row_num, column=time_col, value=current_datetime.strftime("%H:%M:%S"))
        self.ws.cell(row=row_num, column=result_col, value=test_result)

    def find_column_index(self, column_name):
        """Find the column index for a given column name."""
        headers = [cell.value for cell in self.ws[1]]
        for idx, header in enumerate(headers):
            if header == column_name:
                return idx + 1
        return None

    def save_results(self):
        """Save the Excel file with written results."""
        self.wb.save(self.file_path)

    def close(self):
        """Close the workbook."""
        self.wb.close()